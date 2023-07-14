#!/usr/bin/env python3

import pandas as pd
from colorama import Fore
import optparse
import openpyxl
from openpyxl.styles import Font
import re
import csv
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side

csv_list = []
xlsx_file_path = 'output.xlsx'
csv_out = 'csv_out.csv'

def get_arguments():
	parser = optparse.OptionParser()
	parser.add_option("-f", "--file_list", dest="file_list", help="list of nessus file(s) seperated by a comma")
	(options, arguments) = parser.parse_args()
	if not options.file_list:
		print("[" + Fore.RED + "-" + Fore.RESET + "]" + f" Please supply a csv file or list of csv files seperated by a comma")
	return options

def remove_dups(xlsx_file):
	'''
	df = pd.read_excel(xlsx_file, sheet_name = 'Sheet1 Copy')
	df = df.drop_duplicates()
	print("dups should be removed")
	df.to_excel(xlsx_file, sheet_name='Sheet1 Copy', index=False)
'''
	workbook = openpyxl.load_workbook(xlsx_file)

	# Select the specific sheet
	sheet = workbook['Sheet1 Copy']
	# Specify the column indexes to check for duplicates (starting from 1)
	columns_to_check = [3]  # Example: Check columns A, B, and C
	# Create a set to store unique rows based on specified columns
	unique_rows = set()
	rows = list(sheet.iter_rows(min_row=2))

	# Iterate over the rows in reverse order
	for row in reversed(rows):
	    row_values = tuple(cell.value for cell in row)
	    # Generate a key based on column values
	    key = tuple(row_values[column - 1] for column in columns_to_check)
	    # Check if the row is unique based on the key
	    if key in unique_rows:
	        sheet.delete_rows(row[0].row)  # Delete the duplicate row
	    else:
	        unique_rows.add(key)

	# Save both the modified sheet and the entire workbook
	sheet.title = 'Unique Vulns'  # Rename the modified sheet
	#workbook.save('example_modified.xlsx')  # Save the modified sheet
	sheet.auto_filter.ref = sheet.dimensions
	workbook.save(xlsx_file) 

#make the xlsx file
def csv_to_xlsx(csv_file, xlsx_file):
	# Read the CSV file
	df = pd.read_csv(csv_file)
	# Write the DataFrame to an XLSX file
	df.to_excel(xlsx_file, index=False)


#make the top row filterable
def edit_top_row(file_path):
	# Load the workbook

	border_style = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)
	workbook = openpyxl.load_workbook(file_path)
	sheet = workbook.active
	# make the first row bold and size 14 font
	sheet.insert_cols(5)
	cell_col5 = sheet.cell(row=1, column=5)
	cell_col5.value = 'Location'
	cell_col5.alignment = Alignment(horizontal='center', vertical='center')
	sheet.auto_filter.ref = sheet.dimensions
	for cell in sheet[1]:
		column_letter = cell.column_letter
		sheet.column_dimensions[column_letter].width = 25
		cell.font = Font(bold=True, size=14)
	# insert a column next to host
	new_sheet = workbook.copy_worksheet(sheet)
	for cell in new_sheet[1]:
		column_letter = cell.column_letter
		sheet.column_dimensions[column_letter].width = 25
		cell.font = Font(bold=True, size=14)
	col_index = [1, 2, 4, 6, 7, 9, 10, 11, 12, 13]
	col_index.sort(reverse=True)
	for i in col_index:
		new_sheet.delete_cols(i)

	for row in sheet.iter_rows():
		for cell in row:
			cell.border = border_style
	for row in new_sheet.iter_rows():
		for cell in row:
			cell.border = border_style
	sheet.title = 'Vulnerabilities by Host'
	workbook.save(file_path)


def combine_csv_files(csv_list, output_file):
	header = None
	for file in csv_list:
		with open(file, 'r', newline='') as f:
			reader = csv.reader(f)
			header = next(reader)
			break

	with open(output_file, 'w', newline='') as f:
		writer = csv.writer(f)
		writer.writerow(header)


		for file in csv_list:
			with open(file, 'r', newline='') as input_csv:
				reader = csv.reader(input_csv)
				next(reader) 

				for row in reader:
					writer.writerow(row)

def make_list_of_files(file_list):
	csv_list = []
	newfiles = re.sub(r"\s+", "", file_list)
	if "," in newfiles:
		csv_list = newfiles.split(',')
	else:
		csv_list.append(file_list)
	return csv_list

# Convert the CSV file to XLSX

def main():
	options = get_arguments()
	if options.file_list:
		csv_list = make_list_of_files(options.file_list)
	combine_csv_files(csv_list,csv_out)
	csv_to_xlsx('csv_out.csv', xlsx_file_path)
	edit_top_row(xlsx_file_path)
	remove_dups(xlsx_file_path)



if __name__=="__main__":
	main()